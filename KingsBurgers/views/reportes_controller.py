from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv
from datetime import datetime

from KingsBurgers.models import Carrito, CartItem, Cliente, Usuario, Producto

@require_GET
@login_required
def get_carritos_pagados(request):
    """
    Obtiene los carritos pagados con filtros aplicables
    """
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado', 'PAGADO')  # Por defecto solo pagados
    tipo_usuario = request.GET.get('tipo_usuario')
    cliente_id = request.GET.get('cliente_id')
    producto_id = request.GET.get('producto_id')
    buscar = request.GET.get('buscar')

    # Filtrar carritos base
    carritos = Carrito.objects.select_related('usuario').prefetch_related(
        Prefetch('items', queryset=CartItem.objects.select_related('producto'))
    ).order_by('-created_at')

    # Aplicar filtros
    if estado:
        carritos = carritos.filter(estado=estado)
    
    if fecha_inicio and fecha_fin:
        carritos = carritos.filter(created_at__date__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        carritos = carritos.filter(created_at__date__gte=fecha_inicio)
    elif fecha_fin:
        carritos = carritos.filter(created_at__date__lte=fecha_fin)
    
    if tipo_usuario:
        carritos = carritos.filter(usuario__tipo_usuario=tipo_usuario)
    
    if cliente_id:
        carritos = carritos.filter(usuario__cliente__id=cliente_id)
    
    if producto_id:
        carritos = carritos.filter(items__producto__id=producto_id).distinct()
    
    if buscar:
        carritos = carritos.filter(
            Q(usuario__nombre__icontains=buscar) |
            Q(usuario__correo__icontains=buscar) |
            Q(items__producto__nombre__icontains=buscar)
        ).distinct()

    # Paginación manual (si es necesaria para la API)
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 20))
    start = (page - 1) * per_page
    end = start + per_page

    carritos_paginados = carritos[start:end]

    # Construir respuesta
    resultado = []
    for carrito in carritos_paginados:
        try:
            cliente = Cliente.objects.get(usuario=carrito.usuario)
        except Cliente.DoesNotExist:
            cliente = None

        items = []
        for item in carrito.items.all():
            items.append({
                'producto_id': item.producto.id,
                'nombre': item.producto.nombre,
                'precio': float(item.producto.precio),
                'cantidad': item.cantidad,
                'subtotal': float(item.producto.precio * item.cantidad)
            })

        resultado.append({
            'carrito_id': carrito.id,
            'estado': carrito.estado,
            'fecha': carrito.created_at.strftime('%Y-%m-%d %H:%M'),
            'total': sum(item['subtotal'] for item in items),
            'usuario': {
                'id': carrito.usuario.id,
                'nombre': carrito.usuario.nombre,
                'correo': carrito.usuario.correo,
                'tipo_usuario': carrito.usuario.tipo_usuario,
            },
            'cliente': {
                'telefono': cliente.telefono if cliente else None,
                'direccion': cliente.direccion if cliente else None,
            },
            'productos': items,
            'total_items': sum(item['cantidad'] for item in items)
        })

    return JsonResponse({
        'success': True,
        'carritos': resultado,
        'total': carritos.count(),
        'page': page,
        'per_page': per_page
    })

@require_GET
@login_required
def exportar_excel(request):
    """
    Exporta los carritos filtrados a un archivo Excel
    """
    # Reutilizamos la lógica de filtrado de get_carritos_pagados
    carritos = aplicar_filtros_carritos(request)
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Carritos"

    # Estilos para los encabezados
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    # Encabezados
    headers = [
        'ID Carrito', 'Fecha', 'Estado', 
        'ID Usuario', 'Nombre Usuario', 'Correo', 'Tipo Usuario',
        'Teléfono Cliente', 'Dirección Cliente',
        'ID Producto', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # Llenar datos
    row_num = 2
    for carrito in carritos:
        try:
            cliente = Cliente.objects.get(usuario=carrito.usuario)
        except Cliente.DoesNotExist:
            cliente = None

        for item in carrito.items.all():
            ws.cell(row=row_num, column=1, value=carrito.id)
            ws.cell(row=row_num, column=2, value=carrito.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=3, value=carrito.estado)
            
            # Datos usuario
            ws.cell(row=row_num, column=4, value=carrito.usuario.id)
            ws.cell(row=row_num, column=5, value=carrito.usuario.nombre)
            ws.cell(row=row_num, column=6, value=carrito.usuario.correo)
            ws.cell(row=row_num, column=7, value=carrito.usuario.tipo_usuario)
            
            # Datos cliente
            ws.cell(row=row_num, column=8, value=cliente.telefono if cliente else 'N/A')
            ws.cell(row=row_num, column=9, value=cliente.direccion if cliente else 'N/A')
            
            # Datos producto
            ws.cell(row=row_num, column=10, value=item.producto.id)
            ws.cell(row=row_num, column=11, value=item.producto.nombre)
            ws.cell(row=row_num, column=12, value=item.cantidad)
            ws.cell(row=row_num, column=13, value=float(item.producto.precio))
            ws.cell(row=row_num, column=14, value=float(item.producto.precio * item.cantidad))
            
            row_num += 1

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_carritos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

@require_GET
@login_required
def exportar_csv(request):
    """
    Exporta los carritos filtrados a un archivo CSV
    """
    carritos = aplicar_filtros_carritos(request)
    
    response = HttpResponse(content_type='text/csv')
    filename = f"reporte_carritos_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    
    # Encabezados
    headers = [
        'ID Carrito', 'Fecha', 'Estado', 
        'ID Usuario', 'Nombre Usuario', 'Correo', 'Tipo Usuario',
        'Teléfono Cliente', 'Dirección Cliente',
        'ID Producto', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal'
    ]
    writer.writerow(headers)
    
    # Datos
    for carrito in carritos:
        try:
            cliente = Cliente.objects.get(usuario=carrito.usuario)
        except Cliente.DoesNotExist:
            cliente = None

        for item in carrito.items.all():
            row = [
                carrito.id,
                carrito.created_at.strftime('%Y-%m-%d %H:%M'),
                carrito.estado,
                carrito.usuario.id,
                carrito.usuario.nombre,
                carrito.usuario.correo,
                carrito.usuario.tipo_usuario,
                cliente.telefono if cliente else 'N/A',
                cliente.direccion if cliente else 'N/A',
                item.producto.id,
                item.producto.nombre,
                item.cantidad,
                float(item.producto.precio),
                float(item.producto.precio * item.cantidad)
            ]
            writer.writerow(row)
    
    return response

def aplicar_filtros_carritos(request):
    """
    Función auxiliar para aplicar filtros comunes a las consultas de carritos
    """
    # Obtener parámetros de filtro
    carrito_id = request.GET.get('carrito_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado', 'PAGADO')  # Por defecto solo pagados
    tipo_usuario = request.GET.get('tipo_usuario')
    cliente_id = request.GET.get('cliente_id')
    producto_id = request.GET.get('producto_id')
    buscar = request.GET.get('buscar')

    # Filtrar carritos base
    carritos = Carrito.objects.select_related('usuario').prefetch_related(
        Prefetch('items', queryset=CartItem.objects.select_related('producto'))
    ).order_by('-created_at')

    # Aplicar filtros
    if carrito_id:
        carritos = carritos.filter(id=carrito_id)
    
    if estado:
        carritos = carritos.filter(estado=estado)
    
    if fecha_inicio and fecha_fin:
        carritos = carritos.filter(created_at__date__range=[fecha_inicio, fecha_fin])
    elif fecha_inicio:
        carritos = carritos.filter(created_at__date__gte=fecha_inicio)
    elif fecha_fin:
        carritos = carritos.filter(created_at__date__lte=fecha_fin)
    
    if tipo_usuario:
        carritos = carritos.filter(usuario__tipo_usuario=tipo_usuario)
    
    if cliente_id:
        carritos = carritos.filter(usuario__cliente__id=cliente_id)
    
    if producto_id:
        carritos = carritos.filter(items__producto__id=producto_id).distinct()
    
    if buscar:
        carritos = carritos.filter(
            Q(usuario__nombre__icontains=buscar) |
            Q(usuario__correo__icontains=buscar) |
            Q(items__producto__nombre__icontains=buscar)
        ).distinct()

    return carritos

@require_GET
@login_required
def get_filtros_disponibles(request):
    """
    Devuelve los valores disponibles para los filtros
    """
    return JsonResponse({
        'estados': list(Carrito.objects.values_list('estado', flat=True).distinct()),
        'tipos_usuario': list(Usuario.objects.values_list('tipo_usuario', flat=True).distinct()),
        'clientes': list(Cliente.objects.values('id', 'usuario__nombre', 'usuario__correo')),
        'productos': list(Producto.objects.values('id', 'nombre'))
    })